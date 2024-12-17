import openpyxl as op
from openpyxl import Workbook



class CustomerData:

    def __init__(self):
        self.first = None
        self.last = None
        self.addy = None
        self.p_num = None
        self.email = None
        self.cid = None

    def add_customer(self, first, last, addy, p_num, email, cid):
        self.first = first
        self.last = last
        self.addy = addy
        self.p_num = p_num
        self.email = email
        self.cid = int(cid)

        self.wb = op.load_workbook('data.xlsx')
        self.ws = self.wb["customerData"]

        self.ws.append([self.first, self.last, self.addy, self.p_num, self.email, self.cid])
        self.wb.save('data.xlsx')

    def edit_customer(self, new_first, new_last, new_addy, new_phone, new_email, cust_id):
        self.wb = op.load_workbook('data.xlsx')
        self.ws = self.wb["customerData"]
        customer_found = False

        cust_id = int(cust_id)

        for row in self.ws.iter_rows(min_row=2):
            if row[5].value == cust_id:
                customer_found = True
                if new_first is not None:
                    row[0].value = new_first  # Update first name
                if new_last is not None:
                    row[1].value = new_last  # Update last name
                if new_addy is not None:
                    row[2].value = new_addy  # Update address
                if new_phone is not None:
                    row[3].value = new_phone # Update Phonenumber
                if new_email is not None:
                    row[4].value = new_email  # Update email
                break

        if customer_found:
            print("Customer information updated.")
        else:
            print("Customer with ID: ", cust_id, " not found.")

        self.wb.save('data.xlsx')
    
    def get_all_customers(self):
            self.wb = op.load_workbook('data.xlsx')
            self.ws = self.wb["customerData"]
            all_customers = []

            for row in self.ws.iter_rows(min_row=2, values_only=True):
                customer = {
                    "first": row[0],
                    "last": row[1],
                    "addy": row[2],
                    "p_num": row[3],
                    "email": row[4],
                    "cid": row[5]
                }
                all_customers.append(customer)

            return all_customers

    def remove_customer(self, cust_id):
        self.wb = op.load_workbook('data.xlsx')
        self.ws = self.wb["customerData"]
        customer_found = False

        cust_id = int(cust_id)

        for row in self.ws.iter_rows(min_row=2):
            if row[5].value == cust_id:
                customer_found = True
                self.ws.delete_rows(row[0].row)
                print("Customer with ID: " , cust_id, " has been removed.")
                break


        if not customer_found:
            print("Customer with ID: ", cust_id, " not found.")

        self.wb.save('data.xlsx')

class VideoData:

    def __init__(self):
        self.title = None
        self.year = None
        self.director = None
        self.rating = None
        self.genre = None
        self.upc = None
        self.qty = None

    def add_video(self, title, year, director, rating, genre, upc, qty):
        self.title = title
        self.year = year
        self.director = director
        self.rating = rating
        self.genre = genre
        self.upc = int(upc)
        self.qty = int(qty)

        self.wb = op.load_workbook('data.xlsx')
        self.ws = self.wb["videoData"]

        self.ws.append([self.title, self.year, self.director, self.rating, self.genre, self.upc, self.qty])
        self.wb.save('data.xlsx')

    def edit_video(self, new_title, new_year, new_director, new_rating, new_genre, upc_num, new_qty):
        self.wb = op.load_workbook('data.xlsx')
        self.ws = self.wb["videoData"]
        video_found = False
        upc_num = int(upc_num)
        new_qty = int(new_qty)

        for row in self.ws.iter_rows(min_row=1):
            if row[5].value == upc_num:
                video_found = True
                if new_title is not None:
                    row[0].value = new_title
                if new_year is not None:
                    row[1].value = new_year
                if new_director is not None:
                    row[2].value = new_director
                if new_rating is not None:
                    row[3].value = new_rating
                if new_genre is not None:
                    row[4].value = new_genre
                if new_qty is not None:
                    row[6].value = new_qty
                break

        if video_found:
            print("Video information updated.")
        else:
            print("Video with UPC", upc_num, "not found.")

        self.wb.save('data.xlsx')

    def get_all_videos(self):
        self.wb = op.load_workbook('data.xlsx')
        self.ws = self.wb['videoData']
        all_videos = []

        for row in self.ws.iter_rows(min_row=2,values_only=True):
            video = {
                "title":row[0],
                "year":row[1],
                "director":row[2],
                "rating":row[3],
                "genre":row[4],
                "upc":row[5],
                "qty":row[6]
            }
            all_videos.append(video)
        return all_videos

    def remove_video(self, upc_num):
        self.wb = op.load_workbook('data.xlsx')
        self.ws = self.wb["videoData"]
        video_found = False
        upc_num = int(upc_num)

        for row in self.ws.iter_rows(min_row=2):
            if row[5].value == upc_num:
                video_found = True
                self.ws.delete_rows(row[0].row)
                print("Video with UPC", upc_num, "has been removed.")
                break

        if not video_found:
            print("Video with UPC", upc_num, "not found.")

        self.wb.save('data.xlsx')

    def quantity_up(self, upc_num):
        self.wb = op.load_workbook('data.xlsx')
        self.ws = self.wb["videoData"]

        
        upc_num = int(upc_num)

        for row in self.ws.iter_rows(min_row=2):
            if row[5].value == upc_num:
                row[6].value += 1
                break
        self.wb.save('data.xlsx')

    def quantity_down(self, upc_num):
        self.wb = op.load_workbook('data.xlsx')
        self.ws = self.wb["videoData"]
        videofound = True
        
        upc_num = int(upc_num)

        for row in self.ws.iter_rows(min_row=2):
            if row[5].value == upc_num:
                if row[6].value > 0:
                    row[6].value -= 1
                else:
                    videofound = False
                break
        self.wb.save('data.xlsx')
        if videofound is False:
            return False









if __name__ == "__main__":
    #CustomerData().add_customer("tode","henry","this add","123123123","sea mail","123")
    #CustomerData().edit_customer("tode","henry","this add","123123123","sea mail","123")
    #CustomerData().remove_customer("123")
    #print(CustomerData().get_all_customers())
    VideoData().add_video("Robin Hood","1930","charles dwight","5","Lovecraft","88839","9")
    #VideoData().remove_video("88839")



